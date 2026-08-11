from __future__ import annotations

import json
from pathlib import Path

from ..errors import SourceError
from .base import ProductBatch, ProductSource
from .tabular import rows_to_batch


class ExcelProductSource(ProductSource):
    def __init__(self, path: Path):
        self.path = path

    def fetch_products(self) -> ProductBatch:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise SourceError("Excel 업데이트 구성요소를 불러올 수 없습니다.") from exc
        try:
            workbook = load_workbook(self.path, read_only=True, data_only=True)
        except (OSError, ValueError) as exc:
            raise SourceError(f"Excel 파일을 읽을 수 없습니다: {self.path.name}") from exc
        try:
            sheet = (
                workbook["BeyondPack_Master"]
                if "BeyondPack_Master" in workbook.sheetnames
                else workbook.active
            )
            iterator = sheet.iter_rows(values_only=True)
            try:
                headers = [str(value or "").strip().lstrip("\ufeff") for value in next(iterator)]
            except StopIteration as exc:
                raise SourceError("Excel 파일이 비어 있습니다.") from exc
            rows = [dict(zip(headers, values, strict=False)) for values in iterator]
            fingerprint = json.dumps(rows, ensure_ascii=False, default=str).encode("utf-8")
            return rows_to_batch(
                rows,
                source_name=f"Excel({self.path.name})",
                content_fingerprint=fingerprint,
            )
        finally:
            workbook.close()
