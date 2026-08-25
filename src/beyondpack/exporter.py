from __future__ import annotations

from pathlib import Path

from .errors import ConfigurationError
from .packaging import PackagingRepository


HEADERS = [
    "출고건",
    "작업ID",
    "작업일시",
    "작업자",
    "상품DB버전",
    "앱버전",
    "상태",
    "박스그룹ID",
    "시작박스번호",
    "박스수량",
    "무게(kg/BOX)",
    "가로(cm)",
    "세로(cm)",
    "높이(cm)",
    "박스확정시각",
    "FNSKU",
    "품목코드",
    "SKU",
    "국가코드",
    "국가",
    "품목명",
    "상품수량(EA/BOX)",
    "원본수정시각",
]

FIELDS = [
    "shipment_code",
    "job_id",
    "created_at",
    "operator_name",
    "product_db_version",
    "app_version",
    "status",
    "box_group_id",
    "box_start_no",
    "box_count",
    "weight_kg",
    "length_cm",
    "width_cm",
    "height_cm",
    "box_created_at",
    "fnsku",
    "item_code",
    "sku",
    "country_code",
    "country_name",
    "product_name",
    "qty_per_box",
    "source_modified_at",
]


def export_job_xlsx(repo: PackagingRepository, job_id: str, path: Path) -> int:
    return export_rows_xlsx(repo.job_rows(job_id), path)


def export_shipment_xlsx(
    repo: PackagingRepository, shipment_code: str, path: Path
) -> int:
    """출고건 전체를 Excel로 저장한다.

    한 출고건은 날짜와 프로그램 실행을 넘나들며 여러 작업으로 나뉠 수 있으므로,
    실행 단위가 아니라 출고건 단위로 내보내야 실적이 잘리지 않는다.
    """
    return export_rows_xlsx(repo.shipment_rows(shipment_code), path)


def export_rows_xlsx(rows: list[dict], path: Path) -> int:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError as exc:
        raise ConfigurationError("Excel 저장을 위해 openpyxl 패키지를 설치하세요.") from exc

    if not rows:
        raise ValueError("저장된 박스가 없어 Excel로 내보낼 수 없습니다.")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "포장실적"
    sheet.append(HEADERS)
    for row in rows:
        sheet.append([row[field] for field in FIELDS])
    fill = PatternFill("solid", fgColor="0B1F3A")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
    table = Table(displayName="PackagingResults", ref=f"A1:W{len(rows) + 1}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    sheet.add_table(table)
    sheet.freeze_panes = "A2"
    widths = [18, 22, 24, 14, 18, 12, 12, 22, 14, 12, 14, 12, 12, 12, 24, 18, 16, 20, 12, 12, 38, 18, 24]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = width
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return len(rows)
